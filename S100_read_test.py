__auther__ = 'Yoole'

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QPushButton
from PyQt5.QtCore import *
from openpyxl import Workbook
import socket
import time


# #test mode
class S300:
    Voltage = 0.0
    Current = 0.0
    pass


class Ui_MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(612, 494)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.gridLayout = QtWidgets.QGridLayout(self.centralwidget)
        self.gridLayout.setObjectName("gridLayout")

        # file_addr_tag
        self.file_ads_label = QtWidgets.QLabel(self.centralwidget)
        self.file_ads_label.setObjectName("file_ads_label")
        self.gridLayout.addWidget(self.file_ads_label, 0, 1, 1, 1)
        # file_addr_box
        self.file_ads_edit = QtWidgets.QLineEdit(self.centralwidget)
        self.file_ads_edit.setObjectName("file_ads_edit")
        self.gridLayout.addWidget(self.file_ads_edit, 1, 1, 1, 2)

        # ip_port_tag
        self.ip_port_label = QtWidgets.QLabel(self.centralwidget)
        self.ip_port_label.setObjectName("file_ads_label")
        self.gridLayout.addWidget(self.ip_port_label, 2, 1, 1, 1)
        # ip_port_box
        self.ip_edit = QtWidgets.QLineEdit(self.centralwidget)
        self.ip_edit.setObjectName("ip_edit")
        self.gridLayout.addWidget(self.ip_edit, 3, 1, 1, 1)
        self.port_edit = QtWidgets.QLineEdit(self.centralwidget)
        self.port_edit.setObjectName("port_edit")
        self.gridLayout.addWidget(self.port_edit, 3, 2, 1, 1)

        # button
        self.Startbtn = QtWidgets.QPushButton(self.centralwidget)
        self.Startbtn.setObjectName("Startbtn")
        self.Startbtn.setFixedHeight(50)
        self.gridLayout.addWidget(self.Startbtn, 4, 1, 1, 1)
        self.Stopbtn = QtWidgets.QPushButton(self.centralwidget)
        self.Stopbtn.setObjectName("Stopbtn")
        self.Stopbtn.setFixedHeight(50)
        self.gridLayout.addWidget(self.Stopbtn, 4, 2, 1, 1)

        # write_tag
        self.write_label = QtWidgets.QLabel(self.centralwidget)
        self.write_label.setObjectName("write_label")
        self.gridLayout.addWidget(self.write_label, 5, 1, 1, 1)
        # display_box
        self.write_display = QtWidgets.QTextBrowser(self.centralwidget)
        self.write_display.setObjectName("write_display")
        self.gridLayout.addWidget(self.write_display, 6, 1, 1, 2)

        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 612, 26))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.btn1_clicked()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "S300设备测试V1.0"))
        self.file_ads_label.setText(_translate("MainWindow", "文件输出地址："))
        self.ip_port_label.setText(_translate("MainWindow", "IP/ Port："))
        self.Startbtn.setText(_translate("MainWindow", "开始记录"))
        self.Stopbtn.setText(_translate("MainWindow", "停止记录"))
        self.write_label.setText(_translate("MainWindow", "已写入信息："))

    def btn1_clicked(self):

        def cao_start():
            # print('create sheet type:', type(ws))
            fileads_str = self.file_ads_edit.text()
            ip_str = '10.6.15.196'  # self.ip_edit.text()
            port_str = 5025  # self.port_edit.text()
            S300_per = S300()

            # handle info
            ip_port = (ip_str, int(port_str))

            # connect
            cs = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            try:
                err_code = cs.connect_ex(ip_port)
                # print(err_code)
                # print("Connected from", cs.getsockname())
                # print("Connected to", cs.getpeername())
                pass
            except:
                print(err_code)

            wb = Workbook()
            sheet_num = 0
            ws = wb.create_sheet("0", 0)
            ws.append(['电压', '电流'])

            while 1:
                if ws.max_row < 1000 and ws.max_row > 1:
                    ws.append([(S300_per.Voltage), (S300_per.Current)])
                    print(ws.max_row)
                    wb.save(fileads_str + ".xlsx")
                    # clear message
                    S300_per.Voltage = 0.0
                    S300_per.Current = 0.0
                elif ws.max_row == 1000:
                    sheet_num += 1
                    ws = wb.create_sheet(str(sheet_num), sheet_num)
                    ws.append(['电压', '电流'])
                    print('creat new:',ws.max_row)
                    pass


                # 10.6.15.196 5025
                try:
                    time.sleep(1)

                    send = ':READ?'.encode('utf-8')  # *IDN?  :READ?
                    send_data = cs.sendto(send + '\r'.encode() + '\n'.encode(), ip_port)
                    print(send_data)
                    rsv_data = cs.recvfrom(1024)
                    rsv_data_len = len((rsv_data[0].decode()))
                    for index in range(rsv_data_len):
                        if rsv_data[0][index] == 44:
                            field_seperator = index
                        if rsv_data[0][index] == 101:
                            field_seperator_ma = index

                    # unit mA, *10^3
                    # print(type(rsv_data[0][field_seperator + 1:field_seperator_ma]),type(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()))
                    # print(rsv_data[0][field_seperator_ma+2])
                    # print(rsv_data[0][field_seperator_ma + 3])
                    if rsv_data[0][field_seperator_ma+3] == 50:
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()) * 10)
                        # print(S300_per.Current)
                    elif rsv_data[0][field_seperator_ma+3] == 51:
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()))
                        # print(S300_per.Current)
                    elif rsv_data[0][field_seperator_ma+3] == 52:
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()) / 10)
                        # print(S300_per.Current)
                    elif rsv_data[0][field_seperator_ma+3] == 53:
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()) / 100)
                        # print(S300_per.Current)
                    elif rsv_data[0][field_seperator_ma+3] == 54:
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()) / 1000)
                        # print(S300_per.Current)
                    elif rsv_data[0][field_seperator_ma+3] == 55:
                        # print(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode())/(10000))
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()) / (10000))
                        # print((S300_per.Current))
                    elif rsv_data[0][field_seperator_ma+3] == 56:
                        S300_per.Current = '{:.4f}'.format(float(rsv_data[0][field_seperator + 1:field_seperator_ma].decode()) / (100000))
                        # print((S300_per.Current))
                        pass

                    # print(type(rsv_data))
                    # print('receive:', rsv_data[0].decode(), type(rsv_data[0].decode()), rsv_data_len)
                    S300_per.Voltage = '{:.3f}'.format(float(rsv_data[0][0:field_seperator].decode()))
                    # S300_per.Current = rsv_data[0][field_seperator + 1:rsv_data_len].decode()
                    # print('receive-1:', rsv_data[0][0],type(rsv_data[0]))
                    # print('receive-2:', rsv_data[0][1],type(rsv_data[0][0]))
                    print('Voltage-1:', S300_per.Voltage)
                    print('Current-2:', S300_per.Current)

                    # echo and save
                    self.write_display.setText('电压：' + S300_per.Voltage + '\n' \
                                               + '电流：' + S300_per.Current)
                    if ws.max_row == 1:
                        ws.append([(S300_per.Voltage), (S300_per.Current)])
                        wb.save(fileads_str + ".xlsx")

                except Exception as e:
                    print(str(e))

                pass

        self.Startbtn.clicked.connect(cao_start)
        print('dataprocess_btn end.')

# class MyThread(QThread):
#     sec_changed_signal = pyqtSignal(int)
#
#     def __init__(self, sec=1000, parent=None):
#         super().__init__(parent)
#         self.sec = sec
#
#     def run(self):
#         for i in range(self.sec):
#             self.sec_changed_signal.emit(i)
#             time.sleep(1)
#     pass

if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)  # 创建QApplication对象，作为GUI主程序入口

    window = Ui_MainWindow()  # 创建主窗体对象，实例化Ui_MainWindow
    window.show()  # 显示主窗体

    sys.exit(app.exec_())  # 循环中等待退出程序

# import socket
#
# ip_port = ('10.6.15.196', 5025)
#
# cs = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
# # cs.bind(ip_port)
#
# try:
#     err_code = cs.connect_ex(ip_port)
#     print(err_code)
#     print("Connected from", cs.getsockname())
#     print("Connected to", cs.getpeername())
#     pass
# except:
#     print(err_code)
#
# while 1:
#     send = input("cmd:").encode('utf-8') #*IDN?  :READ?
#     send_data = cs.sendto(send+'\r'.encode()+'\n'.encode(), ip_port)
#     print(send_data)
#     rsv_data = cs.recvfrom(1024)
#     print(type(rsv_data))
#     print('receive:', rsv_data[0])
#     pass
