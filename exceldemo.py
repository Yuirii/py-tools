import re
import datetime
import time
import sys
from openpyxl import Workbook
from tkinter import *

# coding:utf-8

# 判断一段文本中是否包含简体中文
# zhmodel = re.compile(u'[\u4e00-\u9fa5]')  #检查中文
zhmodel = re.compile(u'[^\u4e00-\u9fa5]')  # 检查非中文
uu = ["辽宁", "吉林", "黑龙江", "江苏", "浙江", "安徽", "福建", "江西", "山东", "内蒙", "新疆", "西藏", "台湾", "河南", "河北", "山西", "湖北", "湖南", "广东",
      "广西", "海南", "四川", "贵州", "云南", "陕西", "甘肃", "宁夏", "青海", "北京", "上海", "重庆", "天津", "深圳"]


# 检测是否为汉字
def is_Chinese(word):
    match = zhmodel.search(word)
    if match:
        return False
    else:
        return True


# 检测是否为地址
def is_Address(adstr):
    getit = 0
    for m in uu:
        if adstr.find(m) >= 0:
            getit = 1

    if getit == 1:
        return True
    else:
        return False


# 解析文本里，中文名，手机号，和地址，的函数
def extract_info(list_value):
    mobile = str()
    temp = str()
    temp2 = str()
    temp3 = str()
    name = address = ''
    for i in list_value:
        # 第一步先提取手机号
        if i.isnumeric() and len(i) == 11:
            mobile = i
            # 第二步提取中文姓名
        elif (is_Chinese(i)) and (len(i) == 2 or len(i) == 3):
            name = i
            # 第三步分析是否是地址 MC石头CSDN
        elif is_Address(i):
            address = i
            # 剩下的文本全部揉一起
        else:
            temp = ' '.join([temp, i])
    # print('ss1',temp)
    return name, mobile, address, temp


# 准备好Excel
wb = Workbook()  # 创建文件对象
# grab the active worksheet
ws = wb.active  # MC石头CSDN 获取第一个sheet
# 第一行写时间头
# ws.append(['销售整理表', '整理于：', time.strftime("%Y年%m月%d日 %H时%M分%S秒", time.localtime())])  # 写入多个单元格
ws.append(['姓名', '手机', '地址'])
a = datetime.date.today()

# 准备好剪切板监控
r = Tk()
last_string = r.clipboard_get()

#test create excel

while True:
    # 监测频率  MC石头CSDN
    time.sleep(0.3)
    string = r.clipboard_get()
    # 对比内容是否有变
    if string != last_string and string != '':
        address = string
        delivery_address = re.sub('[\s,,]+', ',', address).split(",")
        result = extract_info(delivery_address)

        ws.append([result[0], result[1], result[2], result[3]])

        # 保存到日期命名的Excel
        wb.save("C:/Users/arc/Desktop/婷大大快递信息整理记录 " + a.__format__('%Y-%m-%d') + ".xlsx")
        print('添加excel内容完成：')

        last_string = string

##address = "18119990001  刘上奇     北京市北京市东城区静宁路昌运大厦4楼401号"
##delivery_address = re.sub('[\s,,]+', ',', address).split(",")
##result = extract_info(delivery_address)
##print(result)


### Data can be assigned directly to cells
##ws['A1'] = result[0]      #写入数字
##ws['B1'] = result[1]
##ws['C1'] = result[2]
##ws['E1'] = "你好"+"automation test" #写入中文（unicode中文也可）

### Python types will automatically be converted
##import datetime
##import time
##ws['A2'] = datetime.datetime.now()    #写入一个当前时间
###写入一个自定义的时间格式
##ws['A3'] =time.strftime("%Y年%m月%d日 %H时%M分%S秒",time.localtime())







